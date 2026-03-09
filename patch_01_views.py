"""
Patch payroll/views.py — PAIE-14 (export XLSX) + PAIE-23 (clear_cache).
"""

with open('payroll/views.py', 'r') as f:
    content = f.read()

# ── 1. PAIE-23 : Ajouter action clear_cache sur PayrollParameterViewSet ──
old_param_vs = """class PayrollParameterViewSet(viewsets.ModelViewSet):
    \"\"\"API pour les paramètres de paie.\"\"\"

    queryset = PayrollParameter.objects.all()
    serializer_class = PayrollParameterSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['is_active']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'effective_date']
    ordering = ['code']"""

new_param_vs = """class PayrollParameterViewSet(viewsets.ModelViewSet):
    \"\"\"API pour les parametres de paie.\"\"\"

    queryset = PayrollParameter.objects.all()
    serializer_class = PayrollParameterSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['is_active']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'effective_date']
    ordering = ['code']

    @action(detail=False, methods=['post'])
    def clear_cache(self, request):
        \"\"\"Invalide le cache des parametres de paie.\"\"\"
        from .services.parameter_resolver import PayrollParameterResolver
        PayrollParameterResolver.clear_cache()
        return Response({'success': True, 'message': 'Cache des parametres invalide.'})"""

if old_param_vs in content:
    content = content.replace(old_param_vs, new_param_vs)
    print('OK — action clear_cache ajoutee a PayrollParameterViewSet')
else:
    print('ERREUR — PayrollParameterViewSet introuvable')

# ── 2. PAIE-14 : Ajouter action export_xlsx sur PayrollRunViewSet ──
# Inserer apres generate_summary_pdf
old_summary_end = """    @action(detail=True, methods=['get'])
    def generate_summary_pdf(self, request, pk=None):
        \"\"\"Génère un récapitulatif PDF du lancement de paie.\"\"\"
        payroll_run = self.get_object()

        try:
            pdf_path = PayrollPDFGenerator.generate_payroll_run_summary(payroll_run)
            return Response(
                {
                    'success': True,
                    'message': 'Récapitulatif généré avec succès',
                    'pdf_url': f'/media/{pdf_path}',
                }
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )"""

new_summary_end = (
    old_summary_end
    + """

    @action(detail=True, methods=['get'])
    def export_xlsx(self, request, pk=None):
        \"\"\"Exporte le recapitulatif de paie en XLSX (PAIE-14).\"\"\"
        import io

        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        payroll_run = self.get_object()
        payslips_qs = payroll_run.payslips.all().select_related(
            'employee', 'employee__department'
        ).order_by('employee__last_name', 'employee__first_name')

        if not payslips_qs.exists():
            return Response(
                {'error': 'Aucun bulletin dans ce lancement'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Recapitulatif paie'

        # En-tete
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        headers = [
            'Matricule', 'Nom', 'Prenom', 'Departement',
            'Salaire brut', 'Cotisations sociales', 'Cotisations sante',
            'Impot', 'Net a payer', 'Charges patronales',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Donnees
        num_fmt = '#,##0.00'
        for row_idx, ps in enumerate(payslips_qs, 2):
            ws.cell(row=row_idx, column=1, value=ps.employee.employee_id)
            ws.cell(row=row_idx, column=2, value=ps.employee.last_name)
            ws.cell(row=row_idx, column=3, value=ps.employee.first_name)
            ws.cell(row=row_idx, column=4, value=ps.employee.department.name if ps.employee.department else '')
            for col, val in [
                (5, float(ps.gross_salary)),
                (6, float(ps.cnss_employee)),
                (7, float(ps.amo_employee)),
                (8, float(ps.income_tax)),
                (9, float(ps.net_salary)),
                (10, float(ps.cnss_employer + ps.amo_employer)),
            ]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = num_fmt

        # Totaux
        total_row = payslips_qs.count() + 2
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        for col in range(5, 11):
            letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row, column=col,
                value=f'=SUM({letter}2:{letter}{total_row - 1})',
            )
            cell.font = Font(bold=True)
            cell.number_format = num_fmt

        # Largeurs colonnes
        for col, w in [(1, 12), (2, 18), (3, 15), (4, 18), (5, 15), (6, 16), (7, 16), (8, 14), (9, 15), (10, 17)]:
            ws.column_dimensions[get_column_letter(col)].width = w

        # Reponse HTTP
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        period_name = payroll_run.period.name if payroll_run.period else 'paie'
        filename = f'recapitulatif_paie_{period_name}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response"""
)

if old_summary_end in content:
    content = content.replace(old_summary_end, new_summary_end)
    print('OK — action export_xlsx ajoutee a PayrollRunViewSet')
else:
    print('ERREUR — generate_summary_pdf introuvable')

with open('payroll/views.py', 'w') as f:
    f.write(content)

print('Patch views.py termine.')
