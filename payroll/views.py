import os

from django.db.models import Count, Sum
from django.http import FileResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response

from users.permissions import HasModulePermission, module_permission_required

from .models import (
    AdvanceSalary,
    ContractType,
    EmployeeAllowance,
    EmployeePayroll,
    PayrollParameter,
    PayrollPeriod,
    PayrollRun,
    PaySlip,
    PaySlipLine,
    SalaryComponent,
    TaxBracket,
)
from .serializers import (
    AdvanceSalarySerializer,
    ContractTypeSerializer,
    EmployeeAllowanceSerializer,
    EmployeePayrollSerializer,
    PayrollParameterSerializer,
    PayrollPeriodSerializer,
    PayrollRunSerializer,
    PaySlipLineSerializer,
    PaySlipSerializer,
    SalaryComponentSerializer,
    TaxBracketSerializer,
)
from .services.pdf_generator import PayrollPDFGenerator
from .services.salary_calculator import SalaryCalculator


class PayrollPeriodViewSet(viewsets.ModelViewSet):
    """API pour les périodes de paie."""

    queryset = PayrollPeriod.objects.all()
    serializer_class = PayrollPeriodSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['start_date', 'end_date', 'name']
    ordering = ['-start_date']

    @action(detail=True, methods=['post'])
    def close_period(self, request, pk=None):
        """Clôture une période de paie."""
        period = self.get_object()

        # Vérifier que tous les lancements de paie sont clôturés
        open_runs = PayrollRun.objects.filter(
            period=period, status__in=['draft', 'in_progress', 'calculated']
        )

        if open_runs.exists():
            return Response(
                {
                    'error': 'Impossible de clôturer la période, des lancements de paie sont encore ouverts.'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        period.is_closed = True
        period.save(update_fields=['is_closed'])

        return Response({'success': True, 'message': 'Période clôturée avec succès.'})


class PayrollParameterViewSet(viewsets.ModelViewSet):
    """API pour les parametres de paie."""

    queryset = PayrollParameter.objects.all()
    serializer_class = PayrollParameterSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['is_active']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'effective_date']
    ordering = ['code']

    @action(detail=False, methods=['post'])
    def clear_cache(self, request):
        """Invalide le cache des parametres de paie."""
        from .services.parameter_resolver import PayrollParameterResolver

        PayrollParameterResolver.clear_cache()
        return Response({'success': True, 'message': 'Cache des parametres invalide.'})


class ContractTypeViewSet(viewsets.ModelViewSet):
    """API pour les types de contrat."""

    queryset = ContractType.objects.all()
    serializer_class = ContractTypeSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['is_active']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'name']
    ordering = ['name']


class SalaryComponentViewSet(viewsets.ModelViewSet):
    """API pour les composants de salaire."""

    queryset = SalaryComponent.objects.all()
    serializer_class = SalaryComponentSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['component_type', 'is_taxable', 'is_cnss_eligible', 'is_active']
    search_fields = ['code', 'name', 'description']
    ordering_fields = ['code', 'name', 'component_type']
    ordering = ['component_type', 'code']


class TaxBracketViewSet(viewsets.ModelViewSet):
    """API pour les tranches d'imposition."""

    queryset = TaxBracket.objects.all()
    serializer_class = TaxBracketSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['min_amount', 'effective_date']
    ordering = ['min_amount']


class EmployeePayrollViewSet(viewsets.ModelViewSet):
    """API pour les données de paie des employés."""

    queryset = EmployeePayroll.objects.all()
    serializer_class = EmployeePayrollSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['contract_type', 'payment_method']
    search_fields = [
        'employee__first_name',
        'employee__last_name',
        'employee__employee_id',
        'cnss_number',
        'bank_account',
    ]
    ordering_fields = ['employee__last_name', 'base_salary']
    ordering = ['employee__last_name']

    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Récupère les données de paie pour un employé spécifique."""
        employee_id = request.query_params.get('employee_id')

        if not employee_id:
            return Response(
                {'error': 'employee_id est requis'}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            payroll_info = EmployeePayroll.objects.get(employee__id=employee_id)
            serializer = self.get_serializer(payroll_info)
            return Response(serializer.data)
        except EmployeePayroll.DoesNotExist:
            return Response(
                {'error': 'Informations de paie non trouvées pour cet employé'},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=True, methods=['post'])
    def update_family_status(self, request, pk=None):
        """Met à jour la situation familiale sur le modèle Employee sous-jacent."""
        employee_payroll = self.get_object()
        employee = employee_payroll.employee

        marital_status = request.data.get('marital_status')
        dependent_children = request.data.get('dependent_children')

        update_fields = []
        if marital_status is not None:
            employee.marital_status = marital_status
            update_fields.append('marital_status')
        if dependent_children is not None:
            employee.dependent_children = int(dependent_children)
            update_fields.append('dependent_children')

        if not update_fields:
            return Response(
                {'error': 'Aucune donnée à mettre à jour'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        employee.save(update_fields=update_fields)

        return Response(
            {
                'success': True,
                'message': 'Situation familiale mise à jour',
                'marital_status': employee.marital_status,
                'dependent_children': employee.dependent_children,
            }
        )

    @action(detail=True, methods=['post'])
    def declare_leave_days(self, request, pk=None):
        """Déclare des jours de congé sur le bulletin de la période indiquée."""
        employee_payroll = self.get_object()
        employee = employee_payroll.employee

        period_id = request.data.get('period_id')
        paid_leave_days = request.data.get('paid_leave_days')
        unpaid_leave_days = request.data.get('unpaid_leave_days')

        if not period_id:
            return Response(
                {'error': 'period_id est requis'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payslip = PaySlip.objects.filter(
            payroll_run__period_id=period_id,
            employee=employee,
        ).first()

        if not payslip:
            return Response(
                {'error': 'Aucun bulletin trouvé pour cet employé sur cette période'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if payslip.status not in ('draft', 'calculated'):
            return Response(
                {'error': 'Le bulletin est déjà validé ou payé'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        update_fields = []
        if paid_leave_days is not None:
            from decimal import Decimal as D

            payslip.paid_leave_days = D(str(paid_leave_days))
            update_fields.append('paid_leave_days')
        if unpaid_leave_days is not None:
            from decimal import Decimal as D

            payslip.unpaid_leave_days = D(str(unpaid_leave_days))
            update_fields.append('unpaid_leave_days')

        if update_fields:
            payslip.save(update_fields=update_fields)

        return Response(
            {
                'success': True,
                'message': 'Jours de congé déclarés avec succès',
                'payslip_id': payslip.id,
                'paid_leave_days': str(payslip.paid_leave_days),
                'unpaid_leave_days': str(payslip.unpaid_leave_days),
            }
        )


class EmployeeAllowanceViewSet(viewsets.ModelViewSet):
    """API pour les primes et indemnités dynamiques."""

    queryset = EmployeeAllowance.objects.all()
    serializer_class = EmployeeAllowanceSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
    ]
    filterset_fields = ['employee_payroll', 'component', 'is_active']
    ordering_fields = ['component__code']
    ordering = ['component__code']


class PayrollRunViewSet(viewsets.ModelViewSet):
    """API pour les lancements de paie."""

    queryset = PayrollRun.objects.all()
    serializer_class = PayrollRunSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['period', 'status', 'department']
    search_fields = ['name', 'description']
    ordering_fields = ['period__start_date', 'run_date', 'status']
    ordering = ['-period__start_date', '-run_date']

    def perform_create(self, serializer):
        """Crée le run puis auto-génère les bulletins."""
        import logging

        from hr.models import Employee

        logger = logging.getLogger(__name__)

        try:
            employee = Employee.objects.get(user=self.request.user)
        except Employee.DoesNotExist:
            employee = None

        payroll_run = serializer.save(created_by=employee)

        try:
            count = self._generate_payslips_for_run(payroll_run)
            logger.info(
                'EVO-PAIE-01 : %d bulletins auto-générés pour le run %s',
                count,
                payroll_run.name,
            )
        except Exception as e:
            logger.exception(
                'EVO-PAIE-01 : erreur auto-génération bulletins run %s : %s',
                payroll_run.name,
                e,
            )

    def _generate_payslips_for_run(self, payroll_run):
        """Logique de génération extraite — réutilisée par perform_create et generate_payslips."""
        from decimal import Decimal as D

        from hr.models import Employee, LeaveRequest

        employees = Employee.objects.filter(is_active=True)

        if payroll_run.department:
            employees = employees.filter(department=payroll_run.department)

        period = payroll_run.period
        months = [
            '',
            'JAN',
            'FEV',
            'MAR',
            'AVR',
            'MAI',
            'JUN',
            'JUL',
            'AOU',
            'SEP',
            'OCT',
            'NOV',
            'DEC',
        ]
        month_code = months[period.start_date.month]
        year_code = str(period.start_date.year)[-2:]

        created_count = 0
        for employee in employees:
            try:
                payroll_info = employee.payroll_info
            except EmployeePayroll.DoesNotExist:
                continue

            if PaySlip.objects.filter(
                payroll_run__period=period, employee=employee
            ).exists():
                continue

            number = f'BUL-{month_code}{year_code}-{employee.employee_id}'

            approved_leaves = LeaveRequest.objects.filter(
                employee=employee,
                status='approved_hr',
                start_date__lte=period.end_date,
                end_date__gte=period.start_date,
            ).select_related('leave_type')

            paid_leave_days = sum(
                (req.nb_days for req in approved_leaves if req.leave_type.is_paid),
                D('0'),
            )
            unpaid_leave_days = sum(
                (req.nb_days for req in approved_leaves if not req.leave_type.is_paid),
                D('0'),
            )

            PaySlip.objects.create(
                payroll_run=payroll_run,
                employee=employee,
                number=number,
                basic_salary=payroll_info.base_salary,
                worked_days=26,
                gross_salary=0,
                taxable_salary=0,
                net_salary=0,
                cnss_employee=0,
                cnss_employer=0,
                amo_employee=0,
                amo_employer=0,
                income_tax=0,
                paid_leave_days=paid_leave_days,
                unpaid_leave_days=unpaid_leave_days,
                status='draft',
            )
            created_count += 1

        if created_count > 0:
            payroll_run.status = 'in_progress'
            payroll_run.save(update_fields=['status'])

        return created_count

    @action(detail=True, methods=['post'])
    def generate_payslips(self, request, pk=None):
        """Génère les bulletins de paie pour ce lancement."""
        payroll_run = self.get_object()

        if payroll_run.status != 'draft':
            return Response(
                {
                    'error': "Impossible de générer les bulletins, le lancement n'est pas en brouillon"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = self._generate_payslips_for_run(payroll_run)

        return Response(
            {
                'success': True,
                'message': f'{created_count} bulletins de paie créés avec succès',
                'payslips_count': created_count,
            }
        )

    @action(detail=True, methods=['post'])
    def calculate_payslips(self, request, pk=None):
        """Calcule tous les bulletins de ce lancement."""
        payroll_run = self.get_object()

        if payroll_run.status not in ['in_progress', 'calculated']:
            return Response(
                {
                    'error': "Impossible de calculer les bulletins, le lancement n'est pas en cours ou déjà calculé"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        payslips = PaySlip.objects.filter(payroll_run=payroll_run, status='draft')

        calculated_count = 0
        errors = []

        for payslip in payslips:
            try:
                SalaryCalculator.calculate_payslip(payslip)
                calculated_count += 1
            except Exception as e:
                errors.append(
                    {
                        'payslip_id': payslip.id,
                        'employee': payslip.employee.full_name,
                        'error': str(e),
                    }
                )

        if not PaySlip.objects.filter(payroll_run=payroll_run, status='draft').exists():
            payroll_run.status = 'calculated'
            payroll_run.calculated_date = timezone.now()
            payroll_run.save(update_fields=['status', 'calculated_date'])

        return Response(
            {
                'success': True,
                'message': f'{calculated_count} bulletins calculés avec succès',
                'errors': errors,
            }
        )

    @action(detail=True, methods=['post'])
    def validate_payroll(self, request, pk=None):
        """Valide le lancement de paie."""
        payroll_run = self.get_object()

        if payroll_run.status != 'calculated':
            return Response(
                {
                    'error': 'Impossible de valider le lancement, tous les bulletins ne sont pas calculés'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if (
            PaySlip.objects.filter(payroll_run=payroll_run)
            .exclude(status='calculated')
            .exists()
        ):
            return Response(
                {'error': 'Tous les bulletins ne sont pas calculés'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payroll_run.status = 'validated'
        payroll_run.validated_date = timezone.now()
        try:
            payroll_run.validated_by = request.user.employee
        except Exception:
            payroll_run.validated_by = None
        payroll_run.save(update_fields=['status', 'validated_date', 'validated_by'])

        PaySlip.objects.filter(payroll_run=payroll_run).update(status='validated')

        accounting_error = None
        try:
            from accounting.services.journal_entry_service import JournalEntryService

            entry = JournalEntryService.create_payroll_entry(
                payroll_run, user=request.user
            )
            accounting_message = f'Écriture comptable {entry.name} créée et validée'
        except Exception as e:
            accounting_error = str(e)
            accounting_message = f'Écriture comptable non générée : {accounting_error}'

        return Response(
            {
                'success': True,
                'message': 'Lancement de paie validé avec succès',
                'accounting': accounting_message,
            }
        )

    @action(detail=True, methods=['post'])
    def pay_payroll(self, request, pk=None):
        """Marque le lancement de paie comme payé."""
        payroll_run = self.get_object()

        if payroll_run.status != 'validated':
            return Response(
                {'error': "Impossible de payer le lancement, il n'est pas validé"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payment_reference = request.data.get('payment_reference', '')

        payroll_run.status = 'paid'
        payroll_run.paid_date = timezone.now()
        payroll_run.save(update_fields=['status', 'paid_date'])

        PaySlip.objects.filter(payroll_run=payroll_run).update(
            status='paid',
            is_paid=True,
            payment_date=timezone.now().date(),
            payment_reference=payment_reference,
        )

        return Response(
            {
                'success': True,
                'message': 'Lancement de paie marqué comme payé avec succès',
            }
        )

    @action(detail=True, methods=['get'])
    def generate_summary_pdf(self, request, pk=None):
        """Génère un récapitulatif PDF du lancement de paie."""
        payroll_run = self.get_object()

        try:
            pdf_path = PayrollPDFGenerator.generate_payroll_run_summary(payroll_run)
            return Response(
                {
                    'success': True,
                    'message': 'Récapitulatif généré avec succès',
                    'pdf_url': f'/media/{pdf_path}',
                }
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=['get'])
    def export_xlsx(self, request, pk=None):
        """Exporte le recapitulatif de paie en XLSX (PAIE-14)."""
        import io

        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        payroll_run = self.get_object()
        payslips_qs = (
            payroll_run.payslips.all()
            .select_related('employee', 'employee__department')
            .order_by('employee__last_name', 'employee__first_name')
        )

        if not payslips_qs.exists():
            return Response(
                {'error': 'Aucun bulletin dans ce lancement'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Recapitulatif paie'

        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(
            start_color='1A1A2E', end_color='1A1A2E', fill_type='solid'
        )
        headers = [
            'Matricule',
            'Nom',
            'Prenom',
            'Departement',
            'Salaire brut',
            'Cotisations sociales',
            'Cotisations sante',
            'Impot',
            'Net a payer',
            'Charges patronales',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        num_fmt = '#,##0.00'
        for row_idx, ps in enumerate(payslips_qs, 2):
            ws.cell(row=row_idx, column=1, value=ps.employee.employee_id)
            ws.cell(row=row_idx, column=2, value=ps.employee.last_name)
            ws.cell(row=row_idx, column=3, value=ps.employee.first_name)
            ws.cell(
                row=row_idx,
                column=4,
                value=ps.employee.department.name if ps.employee.department else '',
            )
            for col, val in [
                (5, float(ps.gross_salary)),
                (6, float(ps.cnss_employee)),
                (7, float(ps.amo_employee)),
                (8, float(ps.income_tax)),
                (9, float(ps.net_salary)),
                (10, float(ps.cnss_employer + ps.amo_employer)),
            ]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = num_fmt

        total_row = payslips_qs.count() + 2
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        for col in range(5, 11):
            letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f'=SUM({letter}2:{letter}{total_row - 1})',
            )
            cell.font = Font(bold=True)
            cell.number_format = num_fmt

        for col, w in [
            (1, 12),
            (2, 18),
            (3, 15),
            (4, 18),
            (5, 15),
            (6, 16),
            (7, 16),
            (8, 14),
            (9, 15),
            (10, 17),
        ]:
            ws.column_dimensions[get_column_letter(col)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        period_name = payroll_run.period.name if payroll_run.period else 'paie'
        filename = f'recapitulatif_paie_{period_name}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PaySlipViewSet(viewsets.ModelViewSet):
    """API pour les bulletins de paie."""

    queryset = PaySlip.objects.all()
    serializer_class = PaySlipSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['payroll_run', 'employee', 'status', 'is_paid']
    search_fields = [
        'number',
        'employee__first_name',
        'employee__last_name',
        'employee__employee_id',
    ]
    ordering_fields = [
        'payroll_run__period__start_date',
        'employee__last_name',
        'number',
    ]
    ordering = ['-payroll_run__period__start_date', 'employee__last_name']

    def get_permissions(self):
        """Consultation et téléchargement PDF accessibles à tout employé authentifié."""
        if self.action in ('list', 'retrieve', 'download_pdf'):
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), HasModulePermission()]

    def get_queryset(self):
        """Filtre : employé voit ses bulletins, RH/Finance/superuser voient tout."""
        user = self.request.user
        qs = PaySlip.objects.select_related(
            'payroll_run', 'payroll_run__period', 'employee'
        )
        if user.is_superuser:
            return qs
        try:
            from hr.models import Employee

            emp = Employee.objects.get(user=user)
            if emp.is_hr or emp.is_finance:
                return qs
            return qs.filter(employee=emp)
        except Exception:
            return qs

    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """Calcule ou recalcule un bulletin de paie."""
        payslip = self.get_object()

        if payslip.status not in ['draft', 'calculated']:
            return Response(
                {
                    'error': 'Impossible de calculer ce bulletin, son statut ne le permet pas'
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            SalaryCalculator.calculate_payslip(
                payslip, recalculate=(payslip.status == 'calculated')
            )
            return Response(
                {
                    'success': True,
                    'message': 'Bulletin calculé avec succès',
                    'payslip': PaySlipSerializer(payslip).data,
                }
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """Télécharge le PDF d'un bulletin de paie."""
        payslip = self.get_object()

        if not payslip.pdf_file:
            try:
                pdf_path = PayrollPDFGenerator.generate_payslip_pdf(payslip)
            except Exception as e:
                return Response(
                    {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        from django.conf import settings

        pdf_path = os.path.join(settings.MEDIA_ROOT, payslip.pdf_file)

        if os.path.exists(pdf_path):
            response = FileResponse(
                open(pdf_path, 'rb'), content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="bulletin_{payslip.number}.pdf"'
            )
            return response
        else:
            return Response(
                {'error': "Le fichier PDF n'existe pas"},
                status=status.HTTP_404_NOT_FOUND,
            )


class PaySlipLineViewSet(viewsets.ModelViewSet):
    """API pour les lignes de bulletin de paie."""

    queryset = PaySlipLine.objects.all()
    serializer_class = PaySlipLineSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['payslip', 'component', 'is_employer_contribution']
    ordering_fields = ['display_order', 'component__code']
    ordering = ['display_order', 'component__code']


class AdvanceSalaryViewSet(viewsets.ModelViewSet):
    """API pour les acomptes sur salaire."""

    queryset = AdvanceSalary.objects.all()
    serializer_class = AdvanceSalarySerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'payroll'
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ['employee', 'period', 'is_paid']
    search_fields = ['employee__first_name', 'employee__last_name', 'notes']
    ordering_fields = ['payment_date', 'amount']
    ordering = ['-payment_date']

    @action(detail=True, methods=['post'])
    def mark_as_paid(self, request, pk=None):
        """Marque un acompte comme payé."""
        advance = self.get_object()

        if advance.is_paid:
            return Response(
                {'error': 'Cet acompte est déjà marqué comme payé'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        advance.is_paid = True
        advance.save(update_fields=['is_paid'])

        return Response(
            {'success': True, 'message': 'Acompte marqué comme payé avec succès'}
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def pack_config(request):
    """GET /api/payroll/pack-config/ — Configuration UI complète du pays actif."""
    defaults = {
        'country_code': '',
        'labels': {},
        'form_sections': {
            'classification': {
                'title': 'Classification professionnelle',
                'fields': [
                    {
                        'key': 'professional_category',
                        'label': 'Catégorie professionnelle',
                        'placeholder': 'Ex: Cadre, Agent de maîtrise',
                        'type': 'text',
                        'visible': True,
                    },
                    {
                        'key': 'coefficient',
                        'label': 'Coefficient',
                        'type': 'number',
                        'visible': True,
                    },
                    {
                        'key': 'echelon',
                        'label': 'Échelon',
                        'placeholder': 'Ex: E1, E2',
                        'type': 'text',
                        'visible': True,
                    },
                    {
                        'key': 'indice',
                        'label': 'Indice',
                        'type': 'number',
                        'visible': True,
                    },
                    {
                        'key': 'collective_agreement',
                        'label': 'Convention collective',
                        'placeholder': 'Ex: Convention Collective Interprofessionnelle',
                        'type': 'text',
                        'visible': True,
                    },
                    {
                        'key': 'monthly_hours',
                        'label': 'Horaire mensuel',
                        'placeholder': '173.33',
                        'type': 'number',
                        'visible': True,
                    },
                    {
                        'key': 'igr_parts',
                        'label': 'Parts IGR',
                        'placeholder': '1 à 5',
                        'type': 'number',
                        'visible': False,
                    },
                ],
            },
            'social_ids': {
                'title': 'Identifiants sociaux',
                'fields': [
                    {
                        'key': 'cimr_number',
                        'label': 'N° C.I.M.R',
                        'placeholder': 'Numéro CIMR',
                        'type': 'text',
                        'visible': False,
                    },
                    {
                        'key': 'health_insurance_number',
                        'label': 'N° Assurance Maladie',
                        'placeholder': 'Numéro assurance maladie',
                        'type': 'text',
                        'visible': False,
                    },
                ],
            },
        },
    }
    try:
        from core.models import CompanySetup
        from core.views import COUNTRY_PACKS

        setup = CompanySetup.objects.first()
        if setup and setup.country_code:
            pack = COUNTRY_PACKS.get(setup.country_code, {})
            labels = pack.get('payroll_labels', {})
            form_fields = pack.get('payroll_form_fields', {})

            config = {
                'country_code': setup.country_code,
                'labels': labels,
                'form_sections': defaults['form_sections'],
            }

            for section_key, overrides in form_fields.items():
                if section_key in config['form_sections']:
                    section = config['form_sections'][section_key]
                    for field_override in overrides:
                        for field in section['fields']:
                            if field['key'] == field_override.get('key'):
                                field.update(field_override)
                                break

            return Response(config)
    except Exception:
        pass
    return Response(defaults)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def payroll_labels(request):
    """GET /api/payroll/labels/ — Labels dynamiques selon le pack actif."""
    defaults = {
        'social': 'Cotisations sociales',
        'health': 'Cotisation complémentaire',
        'tax': 'Impôt sur le revenu',
        'social_number': 'N° immatriculation sociale',
        'social_number_short': 'N° immatriculation',
        'social_organism': 'Organisme social',
        'retirement_label': 'Retraite',
        'tax_short': 'Impôt',
    }
    try:
        from core.models import CompanySetup
        from core.views import COUNTRY_PACKS

        setup = CompanySetup.objects.first()
        if setup and setup.country_code:
            labels = dict(
                COUNTRY_PACKS.get(setup.country_code, {}).get(
                    'payroll_labels', defaults
                )
            )
            labels.setdefault(
                'social_employer',
                f'{labels.get("social", "Cotisations sociales")} employeur',
            )
            labels.setdefault(
                'health_employer',
                f'{labels.get("health", "Cotisation complémentaire")} employeur',
            )
            for key, default_val in defaults.items():
                labels.setdefault(key, default_val)
            return Response(labels)
    except Exception:
        pass
    defaults['social_employer'] = 'Cotisations sociales employeur'
    defaults['health_employer'] = 'Cotisation complémentaire employeur'
    return Response(defaults)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
@module_permission_required('payroll')
def payroll_dashboard(request):
    """Tableau de bord du module Paie."""
    current_period = (
        PayrollPeriod.objects.filter(is_closed=False).order_by('-start_date').first()
    )

    payroll_runs = PayrollRun.objects.all()
    runs_by_status = (
        payroll_runs.values('status').annotate(count=Count('id')).order_by('status')
    )

    payslips = PaySlip.objects.all()
    payslips_by_status = (
        payslips.values('status').annotate(count=Count('id')).order_by('status')
    )

    advances = AdvanceSalary.objects.all()
    advances_by_period = (
        advances.values('period__name')
        .annotate(count=Count('id'), total=Sum('amount'))
        .order_by('-period__start_date')[:6]
    )

    if current_period:
        current_period_data = {
            'id': current_period.id,
            'name': current_period.name,
            'start_date': current_period.start_date,
            'end_date': current_period.end_date,
            'runs_count': PayrollRun.objects.filter(period=current_period).count(),
            'payslips_count': PaySlip.objects.filter(
                payroll_run__period=current_period
            ).count(),
            'payslips_calculated': PaySlip.objects.filter(
                payroll_run__period=current_period,
                status__in=['calculated', 'validated', 'paid'],
            ).count(),
            'total_gross': PaySlip.objects.filter(
                payroll_run__period=current_period
            ).aggregate(total=Sum('gross_salary'))['total']
            or 0,
            'total_net': PaySlip.objects.filter(
                payroll_run__period=current_period
            ).aggregate(total=Sum('net_salary'))['total']
            or 0,
        }
    else:
        current_period_data = None

    recent_payslips = PaySlip.objects.all().order_by('-created_at')[:10]
    recent_payslips_data = PaySlipSerializer(recent_payslips, many=True).data

    recent_advances = AdvanceSalary.objects.all().order_by('-created_at')[:10]
    recent_advances_data = AdvanceSalarySerializer(recent_advances, many=True).data

    if current_period:
        department_stats = (
            PaySlip.objects.filter(payroll_run__period=current_period)
            .values('employee__department__name')
            .annotate(
                employees_count=Count('employee', distinct=True),
                total_gross=Sum('gross_salary'),
                total_net=Sum('net_salary'),
                total_cnss=Sum('cnss_employee') + Sum('cnss_employer'),
                total_amo=Sum('amo_employee') + Sum('amo_employer'),
                total_ir=Sum('income_tax'),
            )
            .order_by('-total_gross')
        )
    else:
        department_stats = []

    return Response(
        {
            'current_period': current_period_data,
            'runs_by_status': runs_by_status,
            'payslips_by_status': payslips_by_status,
            'advances_by_period': advances_by_period,
            'recent_payslips': recent_payslips_data,
            'recent_advances': recent_advances_data,
            'department_stats': department_stats,
            'totals': {
                'employees': PaySlip.objects.values('employee').distinct().count(),
                'periods': PayrollPeriod.objects.count(),
                'payroll_runs': payroll_runs.count(),
                'payslips': payslips.count(),
                'advances': advances.count(),
            },
        }
    )
